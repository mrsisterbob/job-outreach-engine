import os
import json
import re
import openpyxl
import requests

EXCEL_PATH = "Job_Outreach_CRM (5).xlsx"
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")

def call_gemini(prompt):
    if not GEMINI_API_KEY:
        print("❌ Error: GEMINI_API_KEY environment variable is missing.")
        return None
    
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3.1-flash-lite:generateContent?key={GEMINI_API_KEY}"
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"response_mime_type": "application/json"}
    }
    
    try:
        res = requests.post(url, json=payload, timeout=15)
        if res.status_code == 200:
            return res.json()["candidates"][0]["content"]["parts"][0]["text"]
    except Exception as e:
        print(f"API Exception: {e}")
    return None

def test_clean_crm():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Could not find {EXCEL_PATH} in the project folder. Make sure it's in the root directory.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    if "Carmen Warm" not in wb.sheetnames:
        print("❌ 'Carmen Warm' sheet not found in workbook.")
        return

    ws = wb["Carmen Warm"]
    
    # Grab the first 10 contact rows (Row 2 to 11)
    rows = list(ws.iter_rows(min_row=2, max_row=11, values_only=False))
    raw_entries = [row[1].value for row in rows if row[1].value]
    
    print(f"🔍 Testing Gemini parsing on the first {len(raw_entries)} contacts...\n")
    
    prompt = f"""You are a CRM data cleaner. Given this list of messy contact strings from a spreadsheet, split each into:
- clean_name: The person's real name (strip out companies or notes).
- company: The official organization/employer name mentioned (e.g. "Phil Santer Spark" -> "Ann Arbor SPARK", "Sam Starks Captrusts" -> "CAPTRUST"). If no company is mentioned, return "".
- note: Any relationship or contextual notes found in the string.

Input list:
{json.dumps(raw_entries, indent=2)}

Respond ONLY with a JSON list matching the input order:
[
  {{"clean_name": "...", "company": "...", "note": "..."}}
]"""

    response_text = call_gemini(prompt)
    if not response_text:
        print("❌ Failed to get response from Gemini API.")
        return

    try:
        cleaned = re.sub(r'^```(?:json)?\s*|\s*```$', '', response_text).strip()
        parsed_results = json.loads(cleaned)
        
        for idx, result in enumerate(parsed_results):
            original = raw_entries[idx]
            print(f"Row {idx + 2}:")
            print(f"  • Original: {original}")
            print(f"  • Cleaned Name: {result.get('clean_name')}")
            print(f"  • Extracted Company: {result.get('company') or '[None Found]'}")
            print(f"  • Extracted Note: {result.get('note') or '[None]'}")
            print("-" * 40)
            
    except Exception as e:
        print(f"❌ JSON Parse Error: {e}\nRaw Response:\n{response_text}")

if __name__ == "__main__":
    test_clean_crm()